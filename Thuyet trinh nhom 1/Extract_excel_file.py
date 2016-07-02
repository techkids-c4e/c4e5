# -*- coding: utf-8 -*-
"""
Created on Sat Jul  2 13:36:42 2016

@author: Hoang To
"""


#tao 1 function tinh luong o day,
# cac bien p (vi tri --> he so luong)
# ngay lam viec (ngay nghi hay ngay thuong)
# gio bat dau lam viec
# gio ket thuc lam viec
#(date.weekday()) Return the day of the week as an integer, where Monday is 0 and Sunday is 6.
from datetime import *
from openpyxl import *
from os import *

def dailysalary(pos,date,checkin,checkout):
    p=0
    a=300000
    if pos=="Chuyên viên":
        p=1
    elif pos=="Trưởng phòng":
        p=1.5
    elif pos=="Giám đốc":
        p=10
    
    if date.weekday()<=4:
        d=1
        s=(checkout-checkin)*p*d*a 
 #co the them truong hop di som, di muon, lam ngoai gio etc vao day
    elif date.weekday()>4:
        d=2
        s=(checkout-checkin)*p*d*a
#ngay nghi luong x2
    return(s)
    
    
#lấy danh sách file trong thư mục
thu_muc=input("Duong dan den thu muc chua file cham cong: ")
file_list=[thu_muc+"\\"+ file for file in os.listdir(thu_muc)]

#Chuan bi file excel
twb=Workbook()
tws=twb.active
tws['A1'].value="Tên nhân viên"
tws['B1'].value="Chức vụ"
tws['C1'].value="Lương tháng"
count=0

#Thực hiện tính toán với từng file
for i in file_list:
#check xem file co phai la file excel
    if i.split(".")[-1]!="xlsx":
        continue
#lấy dữ liệu trên file
    wb=load_workbook(i,data_only=True)
    ws=wb.active
    name=ws.cell('B1').value  #tên người
    pos=ws.cell('B2').value  #Vị trí
    monthsalary=0
#Tính toán
    for row in range(6,ws.max_row):
        date=ws['A'+str(row)].value
        checkin=ws['B'+str(row)].value
        checkout=ws['C'+str(row)].value
        if date==None or checkin==None or checkout==None:
            continue
        monthsalary+=dailysalary(pos,date,checkin,checkout)
        
#ghi dữ liệu sang file excel tổng hợp    
    x=str(tws.max_row+1)
    tws['A'+x]=name
    tws['B'+x]=pos
    tws['C'+x]=monthsalary
    count+=1
    
#Save file
savefile="tonghop.xlsx"
    
twb.save(savefile)

print("Done")
print("Đã tổng hợp ",str(count)," files")
